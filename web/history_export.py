"""
web/history_export.py

Energy Monitor V2

History export helpers for Excel (.xlsx) and PDF.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any
import math
import re
import unicodedata


_INVALID_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")


def safe_filename(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = _INVALID_FILENAME.sub("_", ascii_value).strip("._")
    return cleaned or "history"


def export_filename(device_name: str, extension: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{safe_filename(device_name)}_history_{timestamp}.{extension}"


def _columns(payload: dict[str, Any]) -> list[dict[str, Any]]:
    unit = str(payload.get("unit") or "")

    if payload.get("multi"):
        result = []
        for dataset in payload.get("datasets", []):
            label = str(dataset.get("label") or dataset.get("field") or "Value")
            result.append({
                "title": f"{label} ({unit})" if unit else label,
                "label": label,
                "values": dataset.get("values", []),
            })
        return result

    title = str(payload.get("title") or "Value")
    return [{
        "title": f"{title} ({unit})" if unit else title,
        "label": title,
        "values": payload.get("values", []),
    }]


def _rows(payload: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    labels = payload.get("labels", []) or []
    columns = _columns(payload)
    headers = ["Timestamp"] + [column["title"] for column in columns]
    rows: list[list[Any]] = []

    for index, label in enumerate(labels):
        row: list[Any] = [str(label)]
        for column in columns:
            values = column["values"]
            row.append(values[index] if index < len(values) else None)
        rows.append(row)

    return headers, rows


def _numeric(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def build_xlsx(payload: dict[str, Any]) -> BytesIO:
    """Create a standards-compliant Excel workbook with data and a chart.

    openpyxl is imported lazily so a missing optional Excel dependency does
    not prevent the whole Energy Monitor web application from starting.
    """

    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl. Install it with: "
            "python3 -m pip install openpyxl==3.1.5"
        ) from exc

    headers, rows = _rows(payload)
    device = payload.get("device", {}) or {}
    device_name = str(device.get("name") or "Unknown device")
    title = str(payload.get("title") or "History")
    unit = str(payload.get("unit") or "")
    range_name = str(payload.get("range") or "")
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    decimals = int(payload.get("decimals", 2) or 2)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "History"
    sheet.freeze_panes = "A7"

    blue = "0D6EFD"
    light = "E9ECEF"
    white = "FFFFFF"
    border_color = "D9DEE3"
    thin = Side(style="thin", color=border_color)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    sheet["A1"] = "Energy Monitor V2 - History export"
    sheet["A1"].font = Font(bold=True, size=16, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=blue)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26

    metadata = [
        (2, "Device", device_name),
        (3, "Parameter", f"{title} ({unit})" if unit else title),
        (4, "Range", range_name),
    ]
    for row_number, label, value in metadata:
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        sheet.cell(row_number, 1).fill = PatternFill("solid", fgColor=light)
        sheet.cell(row_number, 1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.cell(row_number, 2).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.cell(4, 3, "Generated")
    sheet.cell(4, 4, generated)
    sheet.cell(4, 3).font = Font(bold=True)
    sheet.cell(4, 3).fill = PatternFill("solid", fgColor=light)
    sheet.cell(4, 3).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.cell(4, 4).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(6, column_index, header)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.row_dimensions[6].height = 22

    for row_index, row in enumerate(rows, start=7):
        sheet.cell(row_index, 1, row[0])
        sheet.cell(row_index, 1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for column_index, value in enumerate(row[1:], start=2):
            number = _numeric(value)
            cell = sheet.cell(row_index, column_index, number if number is not None else "")
            cell.number_format = "0." + ("0" * decimals) if decimals > 0 else "0"
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.column_dimensions["A"].width = 25
    for column_index in range(2, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18
    sheet.auto_filter.ref = f"A6:{get_column_letter(len(headers))}{max(6, 6 + len(rows))}"

    if rows and len(headers) > 1:
        chart = LineChart()
        chart.title = f"{device_name} - {title}"
        chart.style = 13
        chart.y_axis.title = unit or "Value"
        chart.x_axis.title = "Sample"
        chart.height = 9
        chart.width = 19
        data = Reference(sheet, min_col=2, max_col=len(headers), min_row=6, max_row=6 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.legend.position = "b"
        sheet.add_chart(chart, f"{get_column_letter(len(headers) + 2)}2")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _pdf_ascii(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    return text.encode("ascii", "replace").decode("ascii")


def _pdf_escape(value: Any) -> str:
    return _pdf_ascii(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_text(x: float, y: float, text: Any, font: str = "F1", size: float = 8) -> str:
    return f"BT /{font} {size:g} Tf {x:.2f} {y:.2f} Td ({_pdf_escape(text)}) Tj ET\n"


def _truncate(value: Any, max_chars: int) -> str:
    text = _pdf_ascii(value)
    if len(text) <= max_chars:
        return text
    return text[: max(0, max_chars - 3)] + "..."


def _downsample(values: list[float | None], maximum: int = 600) -> list[float | None]:
    if len(values) <= maximum:
        return values
    step = (len(values) - 1) / (maximum - 1)
    return [values[round(index * step)] for index in range(maximum)]


def _chart_page(payload: dict[str, Any], page_width: float, page_height: float) -> bytes:
    device = payload.get("device", {}) or {}
    device_name = str(device.get("name") or "Unknown device")
    title = str(payload.get("title") or "History")
    unit = str(payload.get("unit") or "")
    range_name = str(payload.get("range") or "")
    columns = _columns(payload)

    commands: list[str] = [
        "0.96 0.97 0.98 rg 0 0 842 595 re f\n",
        "1 1 1 rg 24 22 794 551 re f\n",
        "0 0 0 rg\n",
        _pdf_text(34, 548, "Energy Monitor V2 - History export", "F2", 16),
        _pdf_text(34, 529, f"Device: {device_name}", "F1", 9),
        _pdf_text(34, 515, f"Parameter: {title}" + (f" ({unit})" if unit else ""), "F1", 9),
        _pdf_text(34, 501, f"Range: {range_name}", "F1", 9),
        _pdf_text(730, 548, "Graph", "F1", 8),
    ]

    left, bottom, width, height = 72.0, 92.0, 720.0, 365.0
    commands.append(f"0.99 0.99 1 rg {left:.2f} {bottom:.2f} {width:.2f} {height:.2f} re f\n")

    series: list[tuple[str, list[float | None]]] = []
    all_numbers: list[float] = []
    for column in columns:
        values = [_numeric(value) for value in column.get("values", [])]
        values = _downsample(values)
        series.append((str(column.get("label") or column.get("title") or "Value"), values))
        all_numbers.extend(value for value in values if value is not None)

    if not all_numbers:
        commands.append("0 0 0 rg\n")
        commands.append(_pdf_text(left + 260, bottom + 180, "No numeric data in selected range", "F1", 11))
        return "".join(commands).encode("latin-1", "replace")

    minimum = min(all_numbers)
    maximum = max(all_numbers)
    if minimum == maximum:
        padding = max(abs(minimum) * 0.1, 1.0)
        minimum -= padding
        maximum += padding
    else:
        padding = (maximum - minimum) * 0.08
        minimum -= padding
        maximum += padding

    commands.append("0.85 0.87 0.90 RG 0.5 w\n")
    grid_count = 5
    for index in range(grid_count + 1):
        y = bottom + height * index / grid_count
        value = minimum + (maximum - minimum) * index / grid_count
        commands.append(f"{left:.2f} {y:.2f} m {left + width:.2f} {y:.2f} l S\n")
        commands.append("0 0 0 rg\n")
        commands.append(_pdf_text(35, y - 3, f"{value:.2f}", "F1", 7))
        commands.append("0.85 0.87 0.90 RG\n")

    commands.append("0.25 0.28 0.32 RG 0.8 w\n")
    commands.append(f"{left:.2f} {bottom:.2f} m {left:.2f} {bottom + height:.2f} l S\n")
    commands.append(f"{left:.2f} {bottom:.2f} m {left + width:.2f} {bottom:.2f} l S\n")

    colors = [
        (0.05, 0.43, 0.99),
        (0.90, 0.20, 0.20),
        (0.10, 0.60, 0.30),
        (0.95, 0.55, 0.05),
        (0.50, 0.25, 0.75),
    ]

    for series_index, (label, values) in enumerate(series):
        color = colors[series_index % len(colors)]
        commands.append(f"{color[0]} {color[1]} {color[2]} RG 1.4 w\n")
        count = len(values)
        drawing = False
        for index, value in enumerate(values):
            if value is None:
                drawing = False
                continue
            x = left if count <= 1 else left + width * index / (count - 1)
            y = bottom + height * (value - minimum) / (maximum - minimum)
            commands.append(f"{x:.2f} {y:.2f} {'l' if drawing else 'm'}\n")
            drawing = True
        commands.append("S\n")

        legend_x = left + series_index * 145
        legend_y = 65
        commands.append(f"{color[0]} {color[1]} {color[2]} RG 2 w {legend_x:.2f} {legend_y:.2f} m {legend_x + 24:.2f} {legend_y:.2f} l S\n")
        commands.append("0 0 0 rg\n")
        commands.append(_pdf_text(legend_x + 30, legend_y - 3, _truncate(label, 18), "F1", 8))

    commands.append("0 0 0 rg\n")
    commands.append(_pdf_text(left, 44, f"Samples: {len(payload.get('labels', []) or [])}", "F1", 7))
    commands.append(_pdf_text(650, 44, datetime.now().strftime("Generated %Y-%m-%d %H:%M:%S"), "F1", 7))
    return "".join(commands).encode("latin-1", "replace")


def build_pdf(payload: dict[str, Any]) -> BytesIO:
    """Create a landscape PDF containing a graph and the selected history table."""

    headers, rows = _rows(payload)
    device = payload.get("device", {}) or {}
    device_name = str(device.get("name") or "Unknown device")
    title = str(payload.get("title") or "History")
    unit = str(payload.get("unit") or "")
    range_name = str(payload.get("range") or "")

    page_width = 842.0
    page_height = 595.0
    margin = 34.0
    row_height = 13.0
    rows_per_page = 34
    table_width = page_width - 2 * margin
    timestamp_width = 190.0
    remaining = table_width - timestamp_width
    value_width = remaining / max(1, len(headers) - 1)
    widths = [timestamp_width] + [value_width] * (len(headers) - 1)

    pages: list[bytes] = [_chart_page(payload, page_width, page_height)]
    chunks = [rows[i:i + rows_per_page] for i in range(0, len(rows), rows_per_page)] or [[]]
    total_pages = 1 + len(chunks)

    for table_page_number, page_rows in enumerate(chunks, start=1):
        page_number = table_page_number + 1
        commands: list[str] = []
        commands.append("0.96 0.97 0.98 rg 0 0 842 595 re f\n")
        commands.append("1 1 1 rg 24 22 794 551 re f\n")
        commands.append("0 0 0 rg\n")
        commands.append(_pdf_text(margin, 548, "Energy Monitor V2 - History export", "F2", 16))
        commands.append(_pdf_text(margin, 529, f"Device: {device_name}", "F1", 9))
        parameter = f"{title} ({unit})" if unit else title
        commands.append(_pdf_text(margin, 515, f"Parameter: {parameter}", "F1", 9))
        commands.append(_pdf_text(margin, 501, f"Range: {range_name}", "F1", 9))
        commands.append(_pdf_text(720, 548, f"Page {page_number}/{total_pages}", "F1", 8))

        table_top = 477.0
        header_bottom = table_top - 20.0
        commands.append(f"0.05 0.43 0.99 rg {margin:.2f} {header_bottom:.2f} {table_width:.2f} 20 re f\n")
        commands.append("1 1 1 rg\n")

        x = margin
        for index, header in enumerate(headers):
            max_chars = 27 if index == 0 else max(8, int(widths[index] / 5.2))
            commands.append(_pdf_text(x + 4, header_bottom + 6, _truncate(header, max_chars), "F2", 8))
            x += widths[index]

        commands.append("0 0 0 rg 0.78 0.81 0.84 RG 0.5 w\n")
        y = header_bottom
        for row in page_rows:
            next_y = y - row_height
            commands.append(f"{margin:.2f} {next_y:.2f} {table_width:.2f} {row_height:.2f} re S\n")
            x = margin
            for index, value in enumerate(row):
                if index == 0:
                    display = str(value).replace("T", " ").replace("Z", " UTC")
                    max_chars = 31
                else:
                    number = _numeric(value)
                    display = "" if number is None else f"{number:.{int(payload.get('decimals', 2))}f}"
                    max_chars = max(8, int(widths[index] / 5.2))
                commands.append(_pdf_text(x + 4, next_y + 3.5, _truncate(display, max_chars), "F1", 7))
                x += widths[index]
            y = next_y

        x = margin
        total_height = 20 + row_height * len(page_rows)
        for width in widths:
            commands.append(f"{x:.2f} {table_top - total_height:.2f} m {x:.2f} {table_top:.2f} l S\n")
            x += width
        commands.append(f"{x:.2f} {table_top - total_height:.2f} m {x:.2f} {table_top:.2f} l S\n")
        commands.append(_pdf_text(margin, 33, f"Rows in export: {len(rows)}", "F1", 7))
        commands.append(_pdf_text(650, 33, datetime.now().strftime("Generated %Y-%m-%d %H:%M:%S"), "F1", 7))
        pages.append("".join(commands).encode("latin-1", "replace"))

    object_count = 4 + len(pages) * 2
    objects: list[bytes | None] = [None] * (object_count + 1)
    objects[1] = b"<< /Type /Catalog /Pages 2 0 R >>"
    page_ids = [5 + i * 2 for i in range(len(pages))]
    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects[2] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("ascii")
    objects[3] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    objects[4] = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>"

    for index, content in enumerate(pages):
        page_id = 5 + index * 2
        content_id = page_id + 1
        objects[page_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width:g} {page_height:g}] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_id} 0 R >>"
        ).encode("ascii")
        objects[content_id] = (
            f"<< /Length {len(content)} >>\nstream\n".encode("ascii")
            + content
            + b"\nendstream"
        )

    output = BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0] * (object_count + 1)

    for object_id in range(1, object_count + 1):
        offsets[object_id] = output.tell()
        output.write(f"{object_id} 0 obj\n".encode("ascii"))
        output.write(objects[object_id] or b"")
        output.write(b"\nendobj\n")

    xref_offset = output.tell()
    output.write(f"xref\n0 {object_count + 1}\n".encode("ascii"))
    output.write(b"0000000000 65535 f \n")
    for object_id in range(1, object_count + 1):
        output.write(f"{offsets[object_id]:010d} 00000 n \n".encode("ascii"))
    output.write(
        f"trailer\n<< /Size {object_count + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    output.seek(0)
    return output
